import re
import certifi
import urllib3
import datetime
from urllib.parse import urlparse
import smtplib
import ssl
import os
import sys
import pycurl
import validators
from idna import unicode
from openpyxl import Workbook
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

website_mas = []


# проверяем файл и загружаем список сайтов в массив
def upload_website(file_name):
    if os.stat(file_name).st_size:
        global website_mas
        with open(file_name) as i:
            website_mas = i.read().split()
            print('Загружено: %s адресов' % len(website_mas))
    else:
        print('Файл %s пустой!' % file_name)
        print('Работа программы остановлена')
        sys.exit()


# блок тестирования сайта и обработка ошибок
def existing_url(d_name):
    try:
        http = urllib3.PoolManager(cert_reqs='CERT_NONE')
        urllib3.disable_warnings()
        http_code = http.request('POST', d_name, retries=1).status
        return http_code  # == 200
    except urllib3.exceptions.HTTPError as e:
        print(f'WARNING: Этот {d_name} адрес недоступен!')
        return e
    except urllib3.exceptions.NewConnectionError:
        print('Ошибка соединения')
    except Exception:
        print('Непредвиденная ошибка:', sys.exc_info()[0])
        return False


# блок тестирования сайта и обработка ошибок
def check_domain_name(p_url):
    if validators.url(p_url):
        c = pycurl.Curl()
        c.setopt(pycurl.NOBODY, True)
        c.setopt(pycurl.FOLLOWLOCATION, False)
        c.setopt(pycurl.CONNECTTIMEOUT, 10)
        c.setopt(pycurl.TIMEOUT, 10)
        c.setopt(pycurl.COOKIEFILE, '')
        c.setopt(pycurl.URL, p_url)
        c.setopt(pycurl.CAINFO, certifi.where())
        c.setopt(pycurl.VERBOSE, False)
        c.setopt(pycurl.SSL_VERIFYPEER, 1)
        c.setopt(pycurl.SSL_VERIFYHOST, 2)
        try:
            c.perform()
            response_code = c.getinfo(pycurl.RESPONSE_CODE)
            # response_time = c.getinfo(pycurl.TOTAL_TIME)
            # print(response_time)
            c.close()
            # if response_code > 400:
            #     response_code = existing_url(p_url)
            #     # print('Вторая попытка входа была успешной!')
            # return response_code
            # print(response_code)
            return True if response_code < 400 else False
        except pycurl.error as err:
            print('Адрес {} не существует!'.format(p_url))
            print('Код ошибки: {}'.format(err.args[0]))
            print('Описание ошибки: {}'.format(err.args[1]))
            second_check_code = existing_url(p_url)
            if second_check_code == 200:
                print('Вторая проверка ОК, HTTP code: {}'.format(second_check_code))
            else:
                print('Вторая проверка FAILED, HTTP code: {}'.format(second_check_code))
                return False
            return False
    else:
        print('Не верный формат URL адреса!')
        return False


# punycode фильр
def puny_converter(url):
    u_scheme = urlparse(url).scheme
    u_domain = urlparse(url).netloc
    u_path = urlparse(url).path
    puny_name = str(u_domain.encode('idna'), 'utf-8')
    parsed = u_scheme + '://' + puny_name + u_path
    return parsed


upload_website('urls.txt')
#print(website_mas[4])
# url = puny_converter('https://jugfiutfkuy.kjhbkhsdf')  # верный адрес но такого домена не существует
# url = puny_converter('ya.ru')                          # недопустимый URL не указан протокол
# url = puny_converter('https://paycard.beeline.ru')     # проблема с SSL
# url = puny_converter('http://эфг.рф')
# url = puny_converter('https://эфг.рф')

#
# if validators.url(url):
#     print('Нормальный адрес')
# else:
#     print('Неверный адрес')

# print(check_domain_name(url))

# from pycurl import Curl
#
#
# def moodle_smoke_test(url):
#     curl = Curl()
#     curl.setopt(pycurl.URL, url)
#     curl.setopt(pycurl.SSL_VERIFYPEER, False)
#     curl.setopt(pycurl.WRITEFUNCTION, lambda x: None)
#     # curl.perform()
#     # status = curl.getinfo(pycurl.HTTP_CODE)
#     try:
#         curl.perform()
#         status = curl.getinfo(pycurl.HTTP_CODE)
#         print('(ok: {})'.format(status))
#     except pycurl.error as e:
#         print("*** WARNING ***")
#         print('Нет такого домена! : {}'.format(e.args[1]))

    # if status != 200:
    #     print("*** DEPLOY FAILED ***")
    #     print('HTTP Status Code: {}'.format(status))
    #     sys.exit(1)
    # print('(ok: {})'.format(status))


# moodle_smoke_test('https://vk.com/koronapaycom')

# print(url)
# parsed_url = puny_converter(url)
# print(parsed_url)
# print(existing_url(parsed_url))
# print(type(check_domain_name(parsed_url)))

# print(re.findall(r'\d+', )
#
# if check_domain_name(parsed_url) > 400:
#     print(existing_url(parsed_url))

#sys.exit(0)

# инициализация объектов excel
wb = Workbook()
ws = wb.active

# формируем шапку xlsx отчёта
ws.column_dimensions['A'].width = 7
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 18
ws.cell(column=1, row=1).value = "№ п/п"
ws.cell(column=2, row=1).value = "Адрес сайта"
ws.cell(column=3, row=1).value = "Статус"
ws.cell(column=4, row=1).value = "Описание"

out_mass = [[], []]
i = -1

for item in website_mas:
    i += 1
    sys.stdout.flush()
    print("\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b", end='')
    print('Сайтов{0:3} проверено!"'.format(i), end='')
    if check_domain_name(puny_converter(item)):
        ws.cell(column=1, row=i + 2).value = i + 1
        ws.cell(column=2, row=i + 2).value = website_mas[i]
        ws.cell(column=3, row=i + 2).value = 'OK'
        ws.cell(column=4, row=i + 2).value = 'Сайт доступен'
        out_mass[0].append(website_mas[i])
        out_mass[1].append('OK')
        # print(i + 1, '.', mass[i] + ' - Ok')
    else:
        ws.cell(column=1, row=i + 2).value = i + 1
        ws.cell(column=2, row=i + 2).value = website_mas[i]
        ws.cell(column=3, row=i + 2).value = 'FAILED'
        ws.cell(column=4, row=i + 2).value = 'Нет такого сайта!'
        out_mass[0].append(website_mas[i])
        out_mass[1].append('FAILED')
        # print(i + 1, '.', mass[i] + ' - FAILED')

# создаём имена файлов с указанием даты формирования и времени
now = datetime.datetime.now()
cur_datetime = now.strftime('_%d_%m_%Y_%H_%M')
out_filename_txt = 'outfile' + cur_datetime + '.txt'
out_filename_xlsx = 'outfile' + cur_datetime + '.xlsx'
out_filename_json = 'outfile' + cur_datetime + '.json'

# экспорты данных о проверке сайтов в txt файл
# with open(out_filename_txt, 'w') as file_handle:
#     for item in out_mass:
#         file_handle.write('%s\n' % item)

# экспорты данных о проверке сайтов в json файл
# with open(out_filename_json, 'w') as fw:
#     json.dump(out_mass, fw)


# экспорты данных о проверке сайтов в xlsx файл
try:
    wb.save(out_filename_xlsx)
    print('Файл %s выгружен успешно!' % out_filename_xlsx)
except:
    print('Файл %s не выгружен!' % out_filename_xlsx)


# блок отправки отчёта на email с вложенным excel файлом
def email_sending(subject, body, smtp_server, port, sender_email, receiver_email, sender_password):
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = subject
    message["Bcc"] = receiver_email  # Recommended for mass emails
    message.attach(MIMEText(body, "plain"))

    with open(out_filename_xlsx, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())

    encoders.encode_base64(part)

    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {out_filename_xlsx}",
    )

    message.attach(part)
    text = message.as_string()

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, receiver_email, text)
        print(f'Отчёт {out_filename_xlsx} отправлен на Email {sender_email}!')

# email_sending(
#     subject='Отчёт о доступности сайтов',
#     body='Здравствуйте, для Вас сформирован и направлен отчёт о доступности сайтов',
#     smtp_server='smtp.gmail.com',
#     port=465,
#     sender_email='broquemax@gmail.com',
#     receiver_email='m.beloborodov@cft.ru',
#     sender_password='2TfLzo123'
# )
