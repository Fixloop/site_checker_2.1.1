#!/usr/bin/env python
import certifi
import urllib3
import datetime
import smtplib
import pycurl
import ssl
import os
import sys
from urllib.parse import urlparse
from openpyxl import Workbook
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

website_mas = []


# проверяем доступен ли интернет
def check_internet(url='http://google.com'):
    try:
        http = urllib3.PoolManager()
        http.request('GET', url)
        print('Соединение с интернет установлено!')
        return True
    except urllib3.exceptions.HTTPError:
        print('Нет интернета!')
        sys.exit()
        return False


check_internet()


# проверяем файл и загружаем список сайтов в массив
def upload_website(file_name):
    if os.stat(file_name).st_size:
        global website_mas
        with open(file_name) as i:
            website_mas = i.read().split()
            print('Загружено: %s адресов' % len(website_mas))
    else:
        print('Файл %s пустой!' % file_name)
        print('Работа программы остановлена')
        sys.exit()


# блок тестирования сайта и обработка ошибок
def existing_url(d_name):
    try:
        user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) ' \
                     'Chrome/83.0.4103.61 Safari/537.36 '
        http = urllib3.PoolManager(headers=user_agent, cert_reqs='CERT_NONE')
        urllib3.disable_warnings()
        http_code_get = http.request('GET', d_name, retries=2).status
        http_code_post = http.request('POST', d_name, retries=2).status
        # print('\n http_code_get', http_code_get)
        # print('\n http_code_post', http_code_post)
        return True if http_code_get == 200 or http_code_post == 200 else False
        # return True if http_code_post == 200 else False
    except urllib3.exceptions.HTTPError:
        # print(f'\nWARNING: Этот {d_name} адрес недоступен!')
        return False
    except urllib3.exceptions.NewConnectionError:
        # print('\nОшибка соединения')
        return False
    except Exception:
        # print('\nНепредвиденная ошибка:', sys.exc_info()[0])
        return False


def is_url(url):
    return bool(urlparse(url).netloc)


# блок тестирования сайта и обработка ошибок
def check_domain_name(p_url):
    if is_url(p_url):
        c = pycurl.Curl()
        c.setopt(pycurl.USERAGENT, 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, '
                                   'like Gecko) Chrome/39.0.2171.95 Safari/537.36')
        c.setopt(pycurl.NOBODY, True)
        c.setopt(pycurl.FOLLOWLOCATION, True)
        c.setopt(pycurl.CONNECTTIMEOUT, 10)
        c.setopt(pycurl.TIMEOUT, 10)
        c.setopt(pycurl.COOKIEFILE, '')
        c.setopt(pycurl.URL, p_url)
        c.setopt(pycurl.CAINFO, certifi.where())
        c.setopt(pycurl.VERBOSE, False)
        c.setopt(pycurl.SSL_VERIFYPEER, 0)
        c.setopt(pycurl.SSL_VERIFYHOST, 0)
        try:
            c.perform()
            response_code = c.getinfo(pycurl.RESPONSE_CODE)
            c.close()
            if response_code != 200 and existing_url(p_url) != True:
                return False
            else:
                return True
        except pycurl.error as err:
            # print('Адрес {} не существует!, код ошибки {}, описание ошибки:'.format(p_url, err.args[0], err.args[1]))
            second_check_code = existing_url(p_url)
            if second_check_code == 200:
                # print('Вторая проверка ОК, HTTP code: {}'.format(second_check_code))
                return True
            else:
                # print('Вторая проверка FAILED, HTTP code: {}'.format(second_check_code))
                return False
            # return False
    else:
        # print('Не верный формат URL адреса!')
        return False


# punycode фильр
def puny_converter(url):
    scheme, netloc, path, params, query, fragment = urlparse(url)[0:6]
    puny_name = str(netloc.encode('idna'), 'utf-8')
    parsed = scheme + '://' + puny_name + path
    return parsed


upload_website('urls.txt')

wb = Workbook()
ws = wb.active

# формируем шапку xlsx отчёта
ws.column_dimensions['A'].width = 7
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 23
ws.cell(column=1, row=1).value = "№ п/п"
ws.cell(column=2, row=1).value = "Адрес сайта"
ws.cell(column=3, row=1).value = "Статус"
ws.cell(column=4, row=1).value = "Описание"

out_mass = [[], []]
i = -1

for item in website_mas:
    i += 1
    if check_domain_name(puny_converter(item)):
        ws.cell(column=1, row=i + 2).value = i + 1
        ws.cell(column=2, row=i + 2).value = website_mas[i]
        ws.cell(column=3, row=i + 2).value = 'OK'
        ws.cell(column=4, row=i + 2).value = 'Сайт доступен'
        out_mass[0].append(website_mas[i])
        out_mass[1].append('OK')
        # print(i + 1, '.', mass[i] + ' - Ok')
    else:
        ws.cell(column=1, row=i + 2).value = i + 1
        ws.cell(column=2, row=i + 2).value = website_mas[i]
        ws.cell(column=3, row=i + 2).value = 'FAILED'
        ws.cell(column=4, row=i + 2).value = 'Адрес не открывается!'
        out_mass[0].append(website_mas[i])
        out_mass[1].append('FAILED')
        # print(i + 1, '.', mass[i] + ' - FAILED')
    sys.stdout.flush()
    print("\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b\b", end='')
    print('Проверено {0:3} сайтов'.format(i + 1), end='')

# создаём имена файлов с указанием даты формирования и времени
now = datetime.datetime.now()
cur_datetime = now.strftime('_%d_%m_%Y_%H_%M')
out_filename_txt = 'outfile' + cur_datetime + '.txt'
out_filename_xlsx = 'outfile' + cur_datetime + '.xlsx'
out_filename_json = 'outfile' + cur_datetime + '.json'

# экспорты данных о проверке сайтов в txt файл
# with open(out_filename_txt, 'w') as file_handle:
#     for item in out_mass:
#         file_handle.write('%s\n' % item)

# экспорты данных о проверке сайтов в json файл
# with open(out_filename_json, 'w') as fw:
#     json.dump(out_mass, fw)


# экспорты данных о проверке сайтов в xlsx файл
try:
    wb.save(out_filename_xlsx)
    print('\nФайл %s выгружен успешно!' % out_filename_xlsx)
except:
    print('\nФайл %s не выгружен!' % out_filename_xlsx)


# блок отправки отчёта на email с вложенным excel файлом
def email_sending(subject, body, smtp_server, port, sender_email, receiver_email, sender_password):
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = subject
    message["Bcc"] = receiver_email  # Recommended for mass emails
    message.attach(MIMEText(body, "plain"))

    with open(out_filename_xlsx, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())

    encoders.encode_base64(part)

    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {out_filename_xlsx}",
    )

    message.attach(part)
    text = message.as_string()

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, receiver_email, text)
        print(f'Отчёт {out_filename_xlsx} отправлен на Email {sender_email}!')

# email_sending(
#     subject='Отчёт о доступности сайтов',
#     body='Здравствуйте, для Вас сформирован и направлен отчёт о доступности сайтов',
#     smtp_server='smtp.gmail.com',
#     port=465,
#     sender_email='broquemax@gmail.com',
#     receiver_email='m.beloborodov@cft.ru',
#     sender_password='2TfLzo123'
# )
